import openpyxl
import sys

def normalizeSpreadsheet(sourceFileName, sourceSheetName, targetFileName, targetSheetName, sourceCol1, sourceCol2, startRow, endRow, delimiter):
	print("sourceFileName: %s"%sourceFileName)
	print("sourceSheetName: %s"%sourceSheetName)
	print("targetFileName: %s"%targetFileName)
	print("sourceCol1: %s"%sourceCol1)
	print("sourceCol2: %s"%sourceCol2)
	print("startRow: %s"%str(startRow))
	print("endRow: %s"%str(endRow))
	sourceWb = openpyxl.load_workbook(sourceFileName)
	sourceSheet = sourceWb.get_sheet_by_name(sourceSheetName)
	targetWb = openpyxl.Workbook()
	targetWb.create_sheet(targetSheetName)
	targetSheet = targetWb.get_sheet_by_name(targetSheetName)
	targetSheet.cell(row=1, column=1, value=sourceSheet[sourceCol1+"1"].value)
	targetSheet.cell(row=1, column=2, value=sourceSheet[sourceCol2+"1"].value)
	currentTargetRow = 2
	for row in range(int(startRow), int(endRow)):
		print("row: %s"%str(row))
		print("source cell %s"%str(sourceSheet[sourceCol2+str(row)].value))
		if "," in str(sourceSheet[sourceCol2+str(row)].value):
			listToBeFlattened = sourceSheet[sourceCol2+str(row)].value.split(delimiter)
			print("lengthOfList to be flattened: %s" % str(listToBeFlattened))
			lengthOfList = len(listToBeFlattened)
			print("length: %s" % lengthOfList)
			for itemNum in range(0, lengthOfList):
				targetSheet.cell(row=currentTargetRow, column=1, value=sourceSheet[sourceCol1+str(row)].value)
				targetSheet.cell(row=currentTargetRow, column=2, value=listToBeFlattened[itemNum])
				currentTargetRow += 1
		else:
			targetSheet.cell(row=currentTargetRow, column=1, value=sourceSheet[sourceCol1+str(row)].value)
	targetWb.save(targetFileName)
	print("Done")

if __name__ == "__main__":
	normalizeSpreadsheet(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4], sys.argv[5], sys.argv[6], sys.argv[7], sys.argv[8], sys.argv[9])